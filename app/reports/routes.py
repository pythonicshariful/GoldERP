from datetime import date
import calendar
from io import BytesIO
from flask import render_template, request, send_file, flash, redirect, url_for
from flask_login import login_required
from sqlalchemy import func, extract
from ..models import db, GoldSale, Expense, GoldItem, GoldPurchase, MortgageInterest, Cashbook, GoldMortgage
from ..auth.decorators import report_required
from . import reports_bp


def _parse_date_safe(date_str, fallback):
    """Parse a date string, clamping the day to the last valid day of the month
    if the supplied day is out of range (e.g. June 31 → June 30)."""
    try:
        return date.fromisoformat(date_str), date_str
    except ValueError:
        # Try to clamp: split into parts and adjust day
        try:
            parts = date_str.split('-')
            if len(parts) == 3:
                year, month, day = int(parts[0]), int(parts[1]), int(parts[2])
                last_day = calendar.monthrange(year, month)[1]
                clamped = date(year, month, min(day, last_day))
                return clamped, clamped.isoformat()
        except Exception:
            pass
        return fallback, fallback.isoformat()


def _get_date_range():
    default_start = date.today().replace(month=1, day=1)
    default_end = date.today()
    start_str = request.args.get('start_date', default_start.isoformat())
    end_str   = request.args.get('end_date',   default_end.isoformat())
    start, start_str = _parse_date_safe(start_str, default_start)
    end,   end_str   = _parse_date_safe(end_str,   default_end)
    return start, end, start_str, end_str


@reports_bp.route('/pnl')
@login_required
@report_required
def pnl():
    start, end, start_str, end_str = _get_date_range()

    # Revenue: Gold Sales
    sales_revenue = db.session.query(func.sum(GoldSale.paid_amount)).filter(
        GoldSale.sale_date.between(start, end)
    ).scalar() or 0

    making_charges = db.session.query(func.sum(GoldSale.making_charge)).filter(
        GoldSale.sale_date.between(start, end)
    ).scalar() or 0

    # Interest income
    interest_income = db.session.query(func.sum(MortgageInterest.interest_amount)).filter(
        MortgageInterest.interest_month.between(start, end),
        MortgageInterest.received == True
    ).scalar() or 0

    total_revenue = float(sales_revenue) + float(making_charges) + float(interest_income)

    # COGS: estimate from purchases in period (simplified)
    cogs = db.session.query(func.sum(GoldPurchase.total_amount)).filter(
        GoldPurchase.purchase_date.between(start, end),
        GoldPurchase.payment_mode != 'credit'
    ).scalar() or 0

    gross_profit = total_revenue - float(cogs)

    # Expenses by type
    expense_rows = db.session.query(
        Expense.expense_type, func.sum(Expense.amount)
    ).filter(
        Expense.expense_date.between(start, end)
    ).group_by(Expense.expense_type).all()

    total_expenses = sum(float(r[1]) for r in expense_rows)
    net_profit = gross_profit - total_expenses

    return render_template('reports/pnl.html',
                           sales_revenue=float(sales_revenue),
                           making_charges=float(making_charges),
                           interest_income=float(interest_income),
                           total_revenue=total_revenue,
                           cogs=float(cogs),
                           gross_profit=gross_profit,
                           expense_rows=expense_rows,
                           total_expenses=total_expenses,
                           net_profit=net_profit,
                           start_date=start_str, end_date=end_str)


@reports_bp.route('/balance-sheet')
@login_required
@report_required
def balance_sheet():
    # Assets
    cash = db.session.query(func.coalesce(Cashbook.running_balance, 0)).order_by(Cashbook.id.desc()).limit(1).scalar() or 0
    ar = db.session.query(func.sum(GoldSale.balance_due)).filter(GoldSale.balance_due > 0).scalar() or 0
    stock_value = db.session.query(func.sum(GoldItem.current_stock_value)).scalar() or 0
    from ..models import ShowroomSetup
    showroom = ShowroomSetup.query.first()
    advance_rent = float(showroom.advance_rent_paid) if showroom else 0
    decoration = float(showroom.decoration_cost) if showroom else 0

    total_assets = float(cash) + float(ar) + float(stock_value) + advance_rent + decoration

    # Liabilities
    loans_given = db.session.query(func.sum(GoldMortgage.loan_amount)).filter(
        GoldMortgage.status == 'active'
    ).scalar() or 0
    credit_purchases = db.session.query(func.sum(GoldPurchase.total_amount)).filter(
        GoldPurchase.payment_mode == 'credit'
    ).scalar() or 0

    total_liabilities = float(loans_given) + float(credit_purchases)
    equity = total_assets - total_liabilities

    return render_template('reports/balance_sheet.html',
                           cash=float(cash), ar=float(ar),
                           stock_value=float(stock_value),
                           advance_rent=advance_rent, decoration=decoration,
                           total_assets=total_assets,
                           loans_given=float(loans_given),
                           credit_purchases=float(credit_purchases),
                           total_liabilities=total_liabilities,
                           equity=equity)


@reports_bp.route('/cashflow')
@login_required
@report_required
def cashflow():
    start, end, start_str, end_str = _get_date_range()

    receipts = db.session.query(func.sum(Cashbook.amount)).filter(
        Cashbook.txn_date.between(start, end),
        Cashbook.txn_type == 'receipt'
    ).scalar() or 0

    payments = db.session.query(func.sum(Cashbook.amount)).filter(
        Cashbook.txn_date.between(start, end),
        Cashbook.txn_type == 'payment'
    ).scalar() or 0

    net_cash = float(receipts) - float(payments)

    return render_template('reports/cashflow.html',
                           receipts=float(receipts), payments=float(payments),
                           net_cash=net_cash,
                           start_date=start_str, end_date=end_str)


@reports_bp.route('/stock')
@login_required
@report_required
def stock_position():
    category = request.args.get('category', '')
    query = GoldItem.query
    if category:
        query = query.filter(GoldItem.category == category)
    items = query.all()
    total_value = sum(float(i.current_stock_value) for i in items)
    return render_template('reports/stock.html', items=items, total_value=total_value, category=category)


@reports_bp.route('/export/excel')
@login_required
@report_required
def export_excel():
    report_type = request.args.get('type', 'pnl')
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type.upper()

    gold_fill = PatternFill(start_color="B8860B", end_color="B8860B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    if report_type == 'stock':
        items = GoldItem.query.all()
        headers = ['Item Code', 'Item Name', 'Category', 'Unit', 'Stock Qty', 'Stock Value (৳)']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = gold_fill
        for item in items:
            ws.append([item.item_code, item.item_name, item.category, item.unit,
                       float(item.current_stock_qty), float(item.current_stock_value)])
    elif report_type == 'expenses':
        start, end, s, e = _get_date_range()
        expenses = Expense.query.filter(Expense.expense_date.between(start, end)).all()
        headers = ['Date', 'Type', 'Description', 'Amount (৳)', 'Paid By']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = gold_fill
        for exp in expenses:
            ws.append([str(exp.expense_date), exp.expense_type, exp.description,
                       float(exp.amount), exp.paid_by])

    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 4

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'{report_type}_report.xlsx')
